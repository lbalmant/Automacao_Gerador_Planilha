import openpyxl as xl
from openpyxl.styles import PatternFill, Font
import os
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinterdnd2 import DND_FILES, TkinterDnD
import pyautogui as bot
from openpyxl.styles import NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
import sys

class ErroDeDataInvalida:
    pass

def encontrar_caminho_area_de_trabalho():
    # Possíveis caminhos para a Área de Trabalho
    caminhos_possiveis = [
        os.path.join(os.path.expanduser("~"), "Desktop"),
        os.path.join(os.path.expanduser("~"), "Área de Trabalho"),
        os.path.join(os.path.expanduser("~"), "OneDrive", "Área de Trabalho"),
        os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop"),
           
    ]

    # Verifica se algum dos caminhos existe
    for caminho in caminhos_possiveis:
        if os.path.exists(caminho):
            return caminho
    
    # Se nenhum dos caminhos for encontrado, levanta uma exceção
    raise FileNotFoundError("Não foi possível encontrar a pasta Área de Trabalho ou Desktop.")

def encontrar_caminho_relatorio_painel(B, caminhoDesktop, nomepasta):
    #possíveis caminhos para o relatorio painel
    caminhos_possiveis = [
        #caso o governo federal mude o nome das extrações, altere aqui!!!!
        os.path.join(caminhoDesktop, nomepasta, "relatorio_painel ("+str(B)+").xlsx"),
        os.path.join(caminhoDesktop, nomepasta, "relatorio_painel("+str(B)+").xlsx"),
    ]
    #verifica se algum caminho existe
    for caminho in caminhos_possiveis:
        if os.path.exists(caminho):
            return caminho
    else:
        return "e" 

def caminho_relatorio_painel_completo(B, caminho_relatorio):
    caminhos_possiveis = [
        #caso o governo federal mude o nome das extrações, altere aqui!!!!
        os.path.join(caminho_relatorio, "relatorio_painel ("+str(B)+").xlsx"),
        os.path.join(caminho_relatorio, "relatorio_painel("+str(B)+").xlsx"),
    ]    
    for caminho in caminhos_possiveis:
        if os.path.exists(caminho):
            return caminho
    else:
        return "e" 



def adicionar_filtros(planilha):
    #Intervalo para aplicar os filtros
    intervalo_filtro = f"A1:{xl.utils.get_column_letter(planilha.max_column)}1"
    #Filtro
    planilha.auto_filter.ref = intervalo_filtro



def processar_planilhas(event=None):
    try:
        #Remover_border_frame(entrada_nomepasta)
        #Remover_border_frame(entrada_numero_planilhas)
        #Remover_border_frame(entrada_nome_arquivo)
        #verifica_preenchimento()
        nomepasta = entrada_nomepasta.get()
        nomearq = entrada_nome_arquivo.get()
        if(not nomepasta or not nomearq):
            messagebox.showerror("Erro", "Preencha todas as caixas de texto!")
            return
        numeroplanilhas = 300
        verificarEspaco = 0
        for c in nomepasta:
            verificarEspaco += 1


        if nomepasta[verificarEspaco-1]== " ":
            nomepasta = nomepasta[:-1]

        contadorletras = 0
        for a in nomearq:
            contadorletras +=1
            if (a=='/' or a=='\\' or a=='<' or a=='>' or a==':' or a=='|' or a=='?' or a == '*' or a == '.'):
                messagebox.showerror("Erro", "Os nomes de arquivo não podem conter nenhum dos seguintes caracteres: / \\ < > : | ? '' * ")
                return

        
        
        #Testando data
        data_atual_teste = entrada_data.get()
        #método para seguir fluxo normal do programa caso o usuário digite a data no formato DD/MM/AA ao invés de DD/MM/AAAA

        try:
            data_formatada_teste = dt.strptime(data_atual_teste,"%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro", "Data inválida!\nVerifique se a data é valida ou se está digitada no formato DD/MM/AAAA")
            raise ErroDeDataInvalida("Data Inválida fornecida")    
         
        #criando o caminho para a pasta em que foram salvas as planilhas


    
        caminhoDesktop = encontrar_caminho_area_de_trabalho()


        #variável para rodar todas as planilhas
        i=1
        caminhopasta = os.path.join(caminhoDesktop, nomepasta)

        ehCaminhoCompleto = False

        #verifica se a pasta foi encontrada retornando uma mensagem de erro se não for
        if(not os.path.exists(caminhopasta)):
            caminhopasta = nomepasta
            ehCaminhoCompleto = True
        caminho_relatorio = os.path.join(caminhopasta, "relatorio_painel.xlsx")


        if(not os.path.exists(caminhopasta)):
            messagebox.showerror("Pasta não encontrada!","Certifique-se de que a pasta se encontra na área de trabalho do seu Desktop e que o nome esteja correto.\nVoce também pode fornecer o caminho completo para a pasta.")
            return

        while(not os.path.exists(caminho_relatorio) and i<=numeroplanilhas):
            if ehCaminhoCompleto:
                caminho_relatorio = caminho_relatorio_painel_completo(B=i,caminho_relatorio=caminhopasta)
            else:
                caminho_relatorio = encontrar_caminho_relatorio_painel(i, caminhoDesktop, nomepasta)
            i+=1


        #lendo a planilha base
        workbook = xl.load_workbook(caminho_relatorio)
        planilha = workbook.active

        #deletando as 4 primeiras linhas que não fazem parte da nossa análise
        planilha.delete_rows(1,4)

        #workbook.save("exemplo.xlsx")

        #pegar a informação da ultima linha da planilha que estou editando
        UltLinPlanilhaprincipal = planilha.max_row


        #loop que roda todas as planilhas secundarias trazendo as informações importantes para a principal
        while(i!=numeroplanilhas):
            #criando o caminho para cada uma das outras planilhas e ativando a planilha
            CamOutrasPlanilhas = encontrar_caminho_relatorio_painel(i, caminhoDesktop, nomepasta)
            if os.path.exists(CamOutrasPlanilhas):
                Workbook2 = xl.load_workbook(CamOutrasPlanilhas)
                PlanSecundaria = Workbook2.active

                #guardar a informação da última linha para retornar a variavel quando avançar uma coluna
                Linhasprincipaisaux = UltLinPlanilhaprincipal

                colunas = 1
                #rodar todas colunas
                while(colunas <= 12):
                    
                    linhas=6
                    #rodar as linhas
                    while(linhas<=PlanSecundaria.max_row):
                        
                        UltLinPlanilhaprincipal += 1
                        #Capturar informação da planilha secundária para a planilha principal
                        planilha.cell(row=UltLinPlanilhaprincipal, column=colunas).value = PlanSecundaria.cell(row=linhas, column=colunas).value
                        linhas += 1
                    #retorna a contagem das linhas
                    UltLinPlanilhaprincipal = Linhasprincipaisaux
                    colunas += 1
                
                #decremento de 5 unidades pois o planilha secundária tem 5 linhas não utilizadas pela principal
                UltLinPlanilhaprincipal += PlanSecundaria.max_row-5
            i += 1
        planilha.title = ("Planilha Bruta")

        #gerando planilha análise e itens compatíveis
        Planilha_analise(workbook, planilha)
    
        Itens_compativeis(workbook)

        Pesquisa_de_mercado(workbook)

        Racional(workbook=workbook)

        caminho_arquivo = os.path.join(caminhoDesktop, nomepasta, f"{nomearq}.xlsx")
        workbook.save(caminho_arquivo)
    
        messagebox.showinfo("Sucesso", f"Arquivo criado e salvo em: {caminho_arquivo}")
        apagar_caixas_de_texto()
    except FileNotFoundError as e:
        messagebox.showerror("Erro", str(e))
    except ErroDeDataInvalida:
        messagebox.showerror("Erro", "Erro ao processar a data. Execução interrompida.")
    except Exception:
        messagebox.showerror("Erro", "Tente verificar os espaços.")
#salvando o arquivo


def Planilha_analise(workbook, planilha):
        
        #criando a planilha análise
        planilha_analise = workbook.create_sheet(title= "Planilha Análise")
        for row in planilha:
            for cell in row:
                planilha_analise[cell.coordinate].value = cell.value

        j=1

        #preenchimentos para a primeira linha da planilha análise
        fill_cinza = PatternFill(start_color= "CCCCCC", end_color= "CCCCCC", fill_type="solid")
        fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        #criando as tres ultimas colunas da planilha análise
        while(j<16):
            if(j<=12):
                planilha_analise.cell(row=1,column=j).fill = fill_cinza
                j+=1
            else:
                planilha_analise.cell(row = 1, column = j).fill = fill_amarelo
                planilha_analise.cell(row = 1, column=j).font = Font(bold=True)
                j+=1

        planilha_analise.cell(row = 1, column=13).value = "Compatibilidade de descrição"
        planilha_analise.cell(row = 1, column=14).value = "Compatibilidade de unidade de fornecimento"
        planilha_analise.cell(row = 1, column=15).value = "Justificativa"

        #convertendo a coluna 12 para datas e comparação
        data_atual = entrada_data.get()
        #método para seguir fluxo normal do programa caso o usuário digite a data no formato DD/MM/AA ao invés de DD/MM/AAAA
        try:
            data_formatada = dt.strptime(data_atual,"%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro", "Data inválida!\nVerifique se a data é valida ou se está digitada no formato DD/MM/AAAA")
            raise ErroDeDataInvalida("Data Inválida fornecida")

            

        
        data_12_meses_atras = data_formatada - relativedelta(months=12)

        linhas = 2
        money_format = NamedStyle(name='money_format', number_format='"R$ "#,##0.00')

        while(linhas<=planilha_analise.max_row):
            #transformando a coluna j em formato monetario para análise na planilha "Itens compatíveis"
            valor_celula = planilha_analise.cell(row=linhas, column = 8).value
            
            if isinstance(valor_celula,str):
                valor_celula = valor_celula.replace("R$","").strip()
                valor_celula = valor_celula.replace(".", ",").replace(",", ".")
                valor_celula = float(valor_celula)
                
                planilha_analise.cell(row = linhas,column = 8).value = valor_celula
            planilha_analise.cell(row=linhas, column = 8).style = money_format

            #implementação de método para comparação de datas
            valor = planilha_analise.cell(row=linhas, column=12).value
            planilha_analise.cell(row=linhas, column=12).value = dt.strptime(valor, '%d/%m/%Y')
            planilha_analise.cell(row=linhas, column=12).number_format = 'DD/MM/YYYY'
            data_celula = dt.strptime(valor, '%d/%m/%Y')
            
            if (data_celula < data_12_meses_atras):
                planilha_analise.cell(row=linhas, column=13).value = "Não se aplica"
                planilha_analise.cell(row=linhas, column=14).value = "Não se aplica"
                planilha_analise.cell(row=linhas, column=15).value = f"Data anterior ao dia {data_12_meses_atras.strftime('%d/%m/%Y')} (doze meses atrás)"
            linhas+=1  
        #criando filtros
        adicionar_filtros(planilha_analise)
        

def Itens_compativeis(workbook):
    itens_compativeis = workbook.create_sheet(title= "Itens compatíveis") 

    # Aplicar o formato numérico na coluna H da planilha "Itens Compatíveis"

    # Adicionar as fórmulas começando na célula A21
    money_format = NamedStyle(name='money_format', number_format='"R$ "#,##0.00')

    if 'money_format' not in workbook.named_styles:
        workbook.add_named_style(money_format)
    #criando a tabela de analise dos itens compativeis
    itens_compativeis['R2'].value = "MÉDIA"
    itens_compativeis['R3'].value = "DESVIO"
    itens_compativeis['R4'].value = "COEFICIENTE"
    itens_compativeis['R5'].value = "MEDIANA"
    itens_compativeis['R6'].value = '=IF(S4>0.25,"PREÇO MEDIANA","PREÇO MÉDIA")'
    itens_compativeis['R6'].font = Font(bold=True)
    

    
    itens_compativeis['S2'].value = '=AVERAGE(H2:H300)'
    itens_compativeis['S2'].style = money_format

    itens_compativeis['S3'].value = '=STDEVP(H2:H300)'
    itens_compativeis['S3'].style = money_format

    itens_compativeis['S4'].value = '=S3/S2'
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    itens_compativeis['S4'].style = percent_style

    itens_compativeis['S5'].value = '=MEDIAN(H2:H300)'
    itens_compativeis['S5'].style = money_format

    itens_compativeis['S6'].value = '=IF(S4>0.25,S5,S2)'
    itens_compativeis['S6'].style = money_format
    itens_compativeis['S6'].font = Font(bold=True)

    #Aplicando borda na tabelinha
    borda = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for i in range(2,7):
        itens_compativeis.cell(row=i,column=18).border = borda
        itens_compativeis.cell(row= i, column=19).border = borda
            
def Pesquisa_de_mercado(workbook):
    pesquisa_de_mercado = workbook.create_sheet(title= "Pesquisa de mercado")
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')

    money_format = NamedStyle(name='money_format', number_format='"R$ "#,##0.00')
    for i in range(7,17):
        pesquisa_de_mercado.cell(row=i,column = 2).style = money_format
        pesquisa_de_mercado.cell(row=i,column = 3).font = Font(color="0000FF", underline="single")

    pesquisa_de_mercado['A1'].value = "DATA:"
    pesquisa_de_mercado['A2'].value = "HORA:"
    pesquisa_de_mercado['A4'].value = "ESPECIFICAÇÃO:"
    pesquisa_de_mercado['A6'].value = "LOJA"
    pesquisa_de_mercado['B6'].value = "VALOR"
    pesquisa_de_mercado['C6'].value = "SÍTIO"

    pesquisa_de_mercado['A1'].font = Font(bold=True)
    pesquisa_de_mercado['A2'].font = Font(bold=True)
    pesquisa_de_mercado['A4'].font = Font(bold=True)
    pesquisa_de_mercado['A6'].font = Font(bold=True)
    pesquisa_de_mercado['B6'].font = Font(bold=True)
    pesquisa_de_mercado['C6'].font = Font(bold=True)

    pesquisa_de_mercado['A19'].value = 'MÉDIA'
    pesquisa_de_mercado['B19'].value = '=AVERAGE(B7:B16)'
    pesquisa_de_mercado['B19'].style = money_format

    pesquisa_de_mercado['A20'].value = 'DESVIO'
    pesquisa_de_mercado['B20'].value = '=STDEVP(B7:B16)'
    pesquisa_de_mercado['B20'].style = money_format

    pesquisa_de_mercado['A21'].value = 'COEFICIENTE'
    pesquisa_de_mercado['B21'].value = '=B20/B19'
    pesquisa_de_mercado['B20'].style = percent_style

    pesquisa_de_mercado['A22'].value = 'FATOR DE CONVERSÃO'
    pesquisa_de_mercado['B22'].value = 0.634

    pesquisa_de_mercado['A23'].value = "PREÇO FINAL"
    pesquisa_de_mercado['A23'].font = Font(bold=True)

    pesquisa_de_mercado['B23'].value = '=IF(B21>25%,"Erro na Pesquisa",B19*B22)'
    pesquisa_de_mercado['B23'].style = money_format
    borda = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for i in range(19,24):
        pesquisa_de_mercado.cell(row=i,column=1).border = borda
        pesquisa_de_mercado.cell(row= i, column=2).border = borda

def Racional(workbook):
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')

    racional = workbook.create_sheet(title= "Racional")
    racional['A1'].value = "Data análise: "
    racional['A2'].value = "Palavras chaves utilizadas: "
    racional['A3'].value = "O coeficiente de variação encontrado foi de: "
    racional['B3'].value = "=IF(ISBLANK('Itens compatíveis'!S4), 'Pesquisa de mercado'!B21, 'Itens compatíveis'!S4)"
    racional['B3'].style = percent_style
    racional['A5'].value = "Demais considerações:"



def ajuda():
    messagebox.showinfo("Informações importantes",
                        "Recorte as planilhas e cole em uma nova pasta localizada na sua área de trabalho.\n\nCaso a pasta não esteja na sua área de trabalho, você deverá especificar todo o caminho para a pasta.\n\nNão renomeie as planilhas, mantendo os nomes padrões das extrações.\n\nO arquivo final será salvo na mesma pasta onde se encontram as planilhas extraídas.\n\nÉ importante ter um controle dos índices das planilhas salvas na pasta. Não é necessário que estas sigam em sequência, mas o índice de maior número deve ser inferior à 300.")

def apagar_caixas_de_texto():
    entrada_nome_arquivo.delete(0, tk.END)
    entrada_data.delete(0,tk.END)
    entrada_nomepasta.delete(0,tk.END)
    #metodo para fazer o cursor voltar para a primeira caixa
    for i in range (2):
       bot.hotkey('shift', 'tab')
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path,relative_path)

# Criar a janela principal
root = tk.Tk()
root.title("Processador de Planilhas")

#root.attributes("-toolwindow", True)
root.resizable(False,False)


# Layout da interface
tk.Label(root, text="*Nome/Caminho da pasta:").grid(row=0, column=0, padx=10, pady=10)
entrada_nomepasta = tk.Entry(root)
entrada_nomepasta.grid(row=0, column=1, padx=10, pady=10)


tk.Label(root, text="*Nome do arquivo final:").grid(row=1, column=0, padx=10, pady=10)
entrada_nome_arquivo = tk.Entry(root)
entrada_nome_arquivo.grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text = "*Data da extração das planilhas\n(DD/MM/AAAA): ").grid(row = 2, column= 0, padx=10, pady=10)
entrada_data = tk.Entry(root)
entrada_data.grid(row=2, column= 1, padx=10, pady=10)

image_path = resource_path("triangulo.png")
imagem = Image.open(image_path)

imagem = imagem.resize((32,32),Image.LANCZOS)
icone = ImageTk.PhotoImage(imagem)

root.iconphoto(True, icone)

#Criando bota de ajuda
icon_path = resource_path("botao_de_ajuda_transparente.png")
help_icon = Image.open(icon_path)
help_icon = help_icon.resize((22, 22), Image.LANCZOS)

help_icon = ImageTk.PhotoImage(help_icon)

help_button = tk.Button(root, image=help_icon, command=ajuda, borderwidth=0)
help_button.place(x=300, y=170)

# Botão para processar as planilhas
tk.Button(root, text="Processar", command=processar_planilhas).grid(row=4, column=0, columnspan=2, pady=20)
root.bind('<Return>', processar_planilhas)
# Iniciar a aplicação
root.mainloop()